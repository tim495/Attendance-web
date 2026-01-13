from flask import Flask, render_template, request
import pandas as pd
import random
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from datetime import datetime
from openpyxl.utils import get_column_letter
from flask import send_file


app = Flask(__name__)

# 變數命名
today = datetime.now()
today_str = f"{today.year}年{today.month}月{today.day}日"
mouth_str = f"{today.year}年{today.month}月 清洗項目資料夾"

# =====================================================================================
# 原本的 Excel 處理函式與程式碼保持不動
# =====================================================================================

def build_ot_text(row):
    start_str = str(row['起始時間']).split('(')[0].strip()
    end_str = str(row['結束時間']).split('(')[0].strip()
    start_time = pd.to_datetime(start_str).strftime('%H:%M')
    end_time = pd.to_datetime(end_str).strftime('%H:%M')
    hours = float(row['打卡承認時數'])
    hours_str = f"{hours:.1f}" if hours % 1 == 0.5 else str(int(round(hours)))
    return f"{hours_str}H({start_time}-{end_time})"

def is_sunday(date):
    return date.weekday() == 6

def is_saturday(date):
    return date.weekday() == 5

def calc_shift_end_time(date, shift_code):
    base_time = {
        "A7": "16:50",
        "A8": "17:30"
    }.get(shift_code)
    if not base_time:
        return None
    dt = pd.to_datetime(f"{date.date()} {base_time}")
    return (dt + pd.Timedelta(minutes=random.randint(1, 10))).strftime('%H:%M')

def remove_ot_text(original, ot_text):
    return '\n'.join(
        line for line in str(original).splitlines()
        if line.strip() != ot_text
    )

def get_end_time(record):
    lines = str(record).strip().split('\n')
    end_times = []
    for line in lines:
        if '(' in line and '-' in line:
            try:
                end_times.append(line.split('(')[1].split('-')[1].rstrip(')'))
            except:
                pass
    if not end_times:
        return None
    times = [(pd.to_datetime(f"2026-01-01 {t}"), t) for t in end_times]
    return max(times, key=lambda x: x[0])[1] if times else None

# =====================================================================================
# Flask 路由
# =====================================================================================

@app.route("/")
def index():
    return render_template("index.html")  # 預設頁面要有上傳表單

@app.route("/upload", methods=["POST"])
def upload():
    # 取得使用者輸入姓名 & 上傳檔案
    attendance_file = request.files["attendance"]
    overtime_file = request.files["overtime"]

    # 讀取 Excel
    try:
        attendance_df = pd.read_excel(attendance_file, dtype={"員工編號": str}, skiprows=1)
        overtime_df = pd.read_excel(overtime_file, dtype={"員工編號": str})

        attendance_df["日期"] = pd.to_datetime(attendance_df["日期"])
        overtime_df["歸屬日期"] = pd.to_datetime(overtime_df["歸屬日期"])
    except Exception as e:
        print(type(e))
        print(repr(e))
        return render_template("error.html",error_message=f"錯誤型別：{type(e).__name__}，錯誤欄位： {e} ")


    # 篩選員工
    filtered_by_name = attendance_df.copy()

    # 篩選專案加班
    project_overtime_df = overtime_df[
        overtime_df["種類"].astype(str).str.contains("專案", na=False)
       
    ]
    # 只抓有專案加班的出勤資料
    key_df = project_overtime_df[["員工編號", "歸屬日期"]].drop_duplicates()
    filtered_attendance_df = pd.merge(
        attendance_df,
        key_df,
        left_on=["員工編號", "日期"],
        right_on=["員工編號", "歸屬日期"],
        how="inner"
    ).drop(columns=["歸屬日期"])

    # 處理專案加班的時段上下班時間
    for _, ot_row in project_overtime_df.iterrows():
        ot_text = build_ot_text(ot_row)
        mask = filtered_attendance_df["日期"] == ot_row["歸屬日期"]

        if "08:30" in ot_text:
            random_minutes = random.randint(1, 10)
            adjusted_timeA8 = (pd.to_datetime("2026-01-01 08:30") - pd.Timedelta(minutes=random_minutes)).strftime("%H:%M")
            filtered_attendance_df.loc[mask, ["出勤"]] = adjusted_timeA8  

        elif "07:50" in ot_text:
            random_minutes = random.randint(1, 10)
            adjusted_timeA7 = (pd.to_datetime("2026-01-01 07:50") - pd.Timedelta(minutes=random_minutes)).strftime("%H:%M")
            filtered_attendance_df.loc[mask, ["出勤"]] = adjusted_timeA7

        for shift_code in ["A7", "A8"]:
            shift_mask = mask & (filtered_attendance_df["班別代碼"] == shift_code)
            if not shift_mask.any():
                continue
            filtered_attendance_df.loc[shift_mask, "加班紀錄"] = \
                filtered_attendance_df.loc[shift_mask, "加班紀錄"].apply(
                    lambda x: remove_ot_text(x, ot_text)
                )
            weekend = is_sunday(ot_row["歸屬日期"]) or is_saturday(ot_row["歸屬日期"])
            remaining = filtered_attendance_df.loc[shift_mask, "加班紀錄"]
            has_remaining = remaining.apply(lambda x: pd.notna(x) and str(x).strip()).any()

            if ot_text != "1H(12:00)" and weekend and not has_remaining:
                filtered_attendance_df.loc[shift_mask, ["出勤", "Unnamed: 10"]] = ""
            else:
                new_time = calc_shift_end_time(ot_row["歸屬日期"], shift_code)
                if new_time:
                    filtered_attendance_df.loc[shift_mask, "Unnamed: 10"] = new_time

            ot_series = filtered_attendance_df.loc[shift_mask, "加班紀錄"]
            if ot_series.isna().all() or (ot_series.astype(str).str.strip() == "").all():
                filtered_attendance_df.loc[shift_mask, "超時回覆"] = ""

    # 剩餘加班處理（原邏輯不變）
    for idx, row in filtered_attendance_df.iterrows():
        record = row["加班紀錄"]
        date = row["日期"]

        if pd.notna(record) and str(record).strip():
            if is_sunday(date):
                filtered_attendance_df.at[idx, ["出勤", "Unnamed: 10"]] = ""
            else:
                end_str = get_end_time(record)
                if end_str:
                    minutes = int(end_str.split(":")[1])
                    if minutes in (0, 20,30, 50):
                        dt = pd.to_datetime(f"2026-01-01 {end_str}")
                        end_str = (dt + pd.Timedelta(minutes=random.randint(1, 5))).strftime("%H:%M")
                    filtered_attendance_df.at[idx, "Unnamed: 10"] = end_str

    # 合併回所有日期
    final_df = filtered_by_name.set_index(["員工編號", "日期"])
    final_df.update(filtered_attendance_df.set_index(["員工編號", "日期"]))
    final_df = final_df.reset_index()

    # ======================================================
    # 將 Excel 輸出到使用者 OneDrive 資料夾（自動建立）
    # ======================================================
    # user_profile = os.environ.get("USERPROFILE")
    # onedrive_folder = os.path.join(
    #     user_profile,
    #     "OneDrive - 鉅鋼機械股份有限公司 King Steel Machinery Co., Ltd",
    #     "人資共用帳號's files - 考勤記錄清洗資料夾"
    # )
    # print(onedrive_folder)

    # # 最終要存放的資料夾
    # final_folder = os.path.join( onedrive_folder , mouth_str)

    # # 若不存在就建立（存在也不會報錯）
    # os.makedirs(final_folder, exist_ok=True)
    
    # output_path = os.path.join(final_folder, f"清洗後出勤明細 {today_str}.xlsx")
    output_path = os.path.join(f"清洗後出勤明細 {today_str}.xlsx")
    
    final_df = final_df.rename(columns={
    "出勤":"出勤(上班)",
    "Unnamed: 10": "出勤(下班)"
    })  
    final_df = final_df.drop(
    columns=["曠職(時)", "遲到早退", "Unnamed: 20", "Unnamed: 21"],
    errors="ignore"
    )
    
    final_df.to_excel(output_path, index=False)

    # 以下處理 Excel 格式完全不動
    wb = load_workbook(output_path)
    ws = wb.active


  #------------------------------------------------------------------------------------------------------------------------  

    date_col_idx = None
    for idx, col in enumerate(final_df.columns, start=1):
        if col == "日期":
            date_col_idx = idx
            break

    if date_col_idx:
        from openpyxl.styles import numbers
        date_format = "yyyy/m/d"
        for row in range(2, len(final_df)+2):
            cell = ws.cell(row=row, column=date_col_idx)
            if cell.value:
                cell.number_format = date_format

    for col_idx in range(1, len(final_df.columns)+1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 15

    for col_idx, col_name in enumerate(final_df.columns, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        if col_name == "班別時間":
            ws.column_dimensions[col_letter].width = 30
        elif col_name == "差勤紀錄":
            ws.column_dimensions[col_letter].width = 20
        elif col_name == "加班紀錄":
            ws.column_dimensions[col_letter].width = 20

    calibri_font = Font(name='Calibri')
    for row in ws.iter_rows(min_row=1, max_row=len(final_df)+1, min_col=1, max_col=len(final_df.columns)):
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name='Calibri',
                    size=cell.font.size if cell.font.size else 11,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    underline=cell.font.underline,
                    strike=cell.font.strike,
                    color=cell.font.color
                )
            else:
                cell.font = calibri_font
  

    
    table_ref = f"A1:{chr(65 + len(final_df.columns) - 1)}{len(final_df)+1}"
    ws.add_table(Table(displayName="AttendanceTable", ref=table_ref))
    wb.save(output_path )
    
    return send_file(
        output_path,
        as_attachment=True,          # 下載而非在瀏覽器打開
        download_name=f"清洗後出勤明細_{today_str}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



# =====================================================================================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
